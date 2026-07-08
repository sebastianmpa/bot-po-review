"""
florida_outdoor_supplier_service.py
------------------------------------
Implementación concreta de SupplierService para Florida Outdoor Equipment (supplerID = "FO").

CSV esperado por el portal (comma-separated, 3 columnas):
  Item | Quantity | UOM
  - UOM no viene en el JSON → siempre "EA"
  - Solo se procesan los primeros 100 ítems por batch

Precio a comparar: your_price (scraped del carrito, implementación pendiente).
"""

import os
from typing import List, Dict, Optional
from decimal import Decimal

import openpyxl

from models.purchase_model import (
    PurchaseOrderDataModel,
    PurchaseOrderItemModel,
    PurchaseOrderResponseProduct,
    PurchaseOrderResponseData,
)
from services.suppliers.base_supplier_service import SupplierService
from seo_scripts.florida_outdoor_playwright import florida_outdoor_automation_playwright
from seo_scripts.insert_data_in_db import insert_po_review_details, enrich_ltl_from_db
from seo_scripts.rest_consumer_management import register_chunk_item




class FloridaOutdoorSupplierService(SupplierService):
    """
    Estrategia concreta para Florida Outdoor Equipment (FO).
    """

    # ------------------------------------------------------------------ #
    #  Identificación                                                       #
    # ------------------------------------------------------------------ #

    @property
    def supplier_id(self) -> str:
        return "FO"

    @property
    def supplier_name(self) -> str:
        return "Florida Outdoor Equipment"

    # ------------------------------------------------------------------ #
    #  Credenciales                                                         #
    # ------------------------------------------------------------------ #

    def get_credentials(self) -> Dict[str, str]:
        return {
            "email": os.getenv("FOE_USERNAME", "dlrc2652"),
            "password": os.getenv("FOE_PASSWORD", "Elra2026$@"),
        }

    # ------------------------------------------------------------------ #
    #  Formato CSV  (3 columnas, comma-separated)                          #
    # ------------------------------------------------------------------ #

    def csv_headers(self) -> List[str]:
        """Florida Outdoor espera: Item | Quantity | UOM"""
        return ["Item", "Quantity", "UOM"]

    def csv_row(self, product: PurchaseOrderItemModel) -> List:
        """UOM no viene en el JSON — siempre 'EA'."""
        return [product.partNumber, product.qty, "EA"]

    @property
    def _upload_file_extension(self) -> str:
        return ".xlsx"

    def _create_csv(self, products: list, xlsx_filename: str) -> str:
        """
        Override: genera un archivo Excel (.xlsx) en lugar de CSV.
        - Columna Item: formato texto '@' para evitar "number formatted as text"
        - Columna Quantity: entero explícito
        - Limita a los primeros FOE_BATCH_SIZE ítems
        """
        temp_dir = os.path.join(os.path.expanduser("~"), "Downloads", "temp_purchase_orders")
        os.makedirs(temp_dir, exist_ok=True)
        xlsx_path = os.path.join(temp_dir, xlsx_filename)

        print(f"📝 Creando Excel en: {xlsx_path}")
        wb = openpyxl.Workbook()
        ws = wb.active

        # Cabecera
        ws.append(self.csv_headers())

        # Filas de datos — escribir celdas individualmente para forzar tipos
        for product in products:
            row_data = self.csv_row(product)          # [partNumber, qty, "EA"]
            r = ws.max_row + 1

            # Columna 1 (Item): texto explícito con formato '@'
            item_cell = ws.cell(row=r, column=1, value=str(row_data[0]))
            item_cell.number_format = '@'

            # Columna 2 (Quantity): entero
            ws.cell(row=r, column=2, value=int(row_data[1]))

            # Columna 3 (UOM): texto
            ws.cell(row=r, column=3, value=str(row_data[2]))

        wb.save(xlsx_path)
        print(f"✅ Excel creado: {xlsx_filename} ({len(products)} productos)")
        return xlsx_path

    def _cleanup(self, csv_path, final_csv_path) -> None:
        """Limpieza desactivada temporalmente — conservar archivos para inspección."""
        print("⏸️  [Florida Outdoor] Limpieza de archivos desactivada temporalmente.")
        if csv_path:
            print(f"📁 Temp conservado: {csv_path}")
        if final_csv_path:
            print(f"📁 Downloads conservado: {final_csv_path}")

    # ------------------------------------------------------------------ #
    #  Template Method Override: omitir CSV, procesar ítem a ítem          #
    # ------------------------------------------------------------------ #

    def execute(self, po_data: PurchaseOrderDataModel, chunk_id: str) -> PurchaseOrderResponseData:
        """
        Override completo para FOE: después del scraping, cada ítem se procesa,
        inserta en BD y envía al chunk API de forma individual (no en bloque).
        """
        print("=" * 60)
        print(f"🚀 [{self.supplier_name}] PROCESANDO ORDEN DE COMPRA")
        print(f"📦 PO Number: {po_data.poNumber}")
        print(f"🏢 Supplier ID: {po_data.supplerID}")
        print(f"📦 Total de productos: {len(po_data.products)}")
        print("=" * 60)

        try:
            credentials = self.get_credentials()

            # Pre-fetch pack codes antes de la automatización
            pack_map = self._pre_fetch_packs(po_data)
            original_pns_upper = {p.partNumber.upper() for p in po_data.products}
            pack_lookup_index = self._build_pack_lookup_index(pack_map, original_pns_upper)
            extra_dict_items = self._build_pack_extra_dict_items(pack_map, original_pns_upper)

            po_items = [
                {"part_number": p.partNumber, "qty": p.qty, "mfrid": p.mfrid, "mfrid_orig": p.mfrid_orig}
                for p in po_data.products
            ] + extra_dict_items

            # ── Mapas de enriquecimiento (construidos ANTES del scraping) ─────
            ideal_costs: Dict[str, float] = {p.partNumber: p.idealCost for p in po_data.products}
            mfrid_map:   Dict[str, str]   = {p.partNumber: p.mfrid for p in po_data.products}
            mfrid_orig_map: Dict[str, str] = {
                p.partNumber: (p.mfrid_orig or p.mfrid or '') for p in po_data.products
            }

            # ── Pre-fetch LTL: una sola query batch ANTES del scraping ───────
            # Consultamos shipping_ltl para todos los po_items conocidos y
            # construimos un set de PNs con LTL que el callback usará en O(1).
            ltl_check = [
                {'mfrid': it.get('mfrid', ''), 'part_number': it.get('part_number', '')}
                for it in po_items if it.get('part_number')
            ]
            enrich_ltl_from_db(ltl_check)
            ltl_pns_set = {it['part_number'].upper() for it in ltl_check if it.get('ltl') == 'Y'}
            if ltl_pns_set:
                print(f"  🚛 {len(ltl_pns_set)} ítem(s) con LTL pre-fetched desde shipping_ltl.")
            else:
                print("  ℹ️  Sin LTL en shipping_ltl para esta PO.")

            total_items = len(po_data.products)
            item_counter = [0]   # lista mutable — permite mutación desde la closure
            response_products: List[PurchaseOrderResponseProduct] = []

            # ── Callback: invocado por cada ítem en cuanto su stock es scrapeado
            def on_item_ready(item: dict) -> None:
                part = item.get('part_number', '')
                if not part:
                    return
                # Saltar ítems de pack (consultas de precio auxiliares, no PO reales)
                if part.upper() in pack_lookup_index:
                    return

                item_counter[0] += 1
                i = item_counter[0]

                # Enriquecer metadata
                item['po_number']     = po_data.poNumber
                item['supplier_code'] = po_data.supplerID
                if not item.get('mfrid'):
                    item['mfrid'] = mfrid_map.get(part, '')
                item['mfrid_orig'] = item.get('mfrid_orig') or mfrid_orig_map.get(part, '')
                if item.get('status') == 'SUPERSEDED' and item.get('superseded_from'):
                    item['partnumber_orig'] = item['superseded_from']
                else:
                    item['partnumber_orig'] = item.get('partnumber_orig') or part
                item['ideal_cost'] = ideal_costs.get(part, 0.0)

                # LTL desde pre-fetched set (O(1), sin query adicional por ítem)
                if not item.get('ltl') and part.upper() in ltl_pns_set:
                    item['ltl'] = 'Y'

                # Comparación de precios
                your_price      = item.get('your_price')
                price_float     = float(your_price) if your_price is not None else 0.0
                ideal_cost_v    = item.get('ideal_cost', 0.0)
                pre_status      = item.get('status', 'CORRECT')
                err_msg         = item.get('error_message')
                pack_qty        = item.get('pack_qty')
                ltl             = item.get('ltl')
                superseded_from = item.get('superseded_from')
                cart_qty        = item.get('qty', 0)

                if pre_status == 'PART_ERROR' and price_float == 0.0:
                    status = 'PART_ERROR'
                    print(f"  [{i}/{total_items}] ❌ PART_ERROR [{part}]: {err_msg}")
                elif pre_status == 'SUPERSEDED':
                    status = 'SUPERSEDED'
                    print(f"  [{i}/{total_items}] 🔄 SUPERSEDED [{part}] ← {superseded_from}")
                elif price_float == 0.0:
                    status = 'CORRECT'
                    print(f"  [{i}/{total_items}] ⏳ SIN PRECIO: {part}")
                elif price_float > 0 and ideal_cost_v > 0:
                    difference = abs(ideal_cost_v - price_float)
                    tolerance  = ideal_cost_v * 0.01
                    print(
                        f"  [{i}/{total_items}] 💵 [{part}] "
                        f"Ideal=${ideal_cost_v:.2f} | FOE=${price_float:.2f} | "
                        f"Diff=${difference:.2f} | Tol=${tolerance:.2f}"
                    )
                    if difference > tolerance:
                        status = 'MISMATCH'
                        err_msg = (
                            f"Price mismatch: Expected ${ideal_cost_v:.2f}, FOE ${price_float:.2f}"
                        )
                        if pack_qty:
                            err_msg += f" (Pack item — min qty: {pack_qty})"
                        print(f"  [{i}/{total_items}] ❌ MISMATCH")
                    else:
                        status = 'CORRECT'
                        notes = []
                        if pack_qty:
                            notes.append(f"Pack item — min qty: {pack_qty}")
                        if ltl:
                            notes.append("LTL shipment required")
                        err_msg = " | ".join(notes) if notes else None
                        print(f"  [{i}/{total_items}] ✅ CORRECT")
                else:
                    status = 'CORRECT'

                item['status']        = status
                item['error_message'] = err_msg

                final_mfrid = item.get('mfrid') or mfrid_map.get(part, '')
                product = PurchaseOrderResponseProduct(
                    mfrid=final_mfrid,
                    partNumber=part,
                    qty=cart_qty,
                    idealCost=ideal_cost_v if ideal_cost_v > 0 else 0.0,
                    supplierPrice=price_float,
                    status=status,
                    nla=item.get('nla'),
                    supersededFrom=superseded_from,
                    packQty=pack_qty,
                    ltl=ltl,
                )
                response_products.append(product)

                # 1. Insertar en BD inmediatamente
                try:
                    insert_po_review_details([item])
                    print(f"  [{i}/{total_items}] 💾 BD: {part}")
                except Exception as db_err:
                    print(f"  [{i}/{total_items}] ⚠️ Error BD [{part}]: {db_err}")

                # 2. Enviar chunk progress inmediatamente
                try:
                    register_chunk_item({
                        "chunkId": chunk_id,
                        "item": {
                            "poNumber":  po_data.poNumber,
                            "supplerID": po_data.supplerID,
                            "products":  [product.dict()],
                        },
                        "status": "Success",
                    })
                    print(f"  [{i}/{total_items}] 📤 Chunk: {part}")
                except Exception as chunk_err:
                    print(f"  [{i}/{total_items}] ⚠️ Error chunk [{part}]: {chunk_err}")

            # ── Ejecutar automatización — el callback se invoca por cada ítem ──
            print(f"🤖 Ejecutando automatización [{self.supplier_name}]...")
            scraped_data = self.run_automation(
                email=credentials["email"],
                password=credentials["password"],
                csv_filename="",
                po_data=po_data,
                po_items=po_items,
                on_item_ready=on_item_ready,
            )

            if not scraped_data:
                raise RuntimeError(
                    f"[{self.supplier_name}] No se obtuvieron datos del scraping "
                    f"para PO {po_data.poNumber}"
                )

            # Pack cost extraction (extrae precios de ítems auxiliares de pack)
            if pack_lookup_index:
                scraped_data = self._apply_pack_costs_and_clean(
                    scraped_data, pack_map, pack_lookup_index
                )

            self._print_summary(response_products)
            return PurchaseOrderResponseData(
                poNumber=po_data.poNumber,
                supplerID=po_data.supplerID,
                products=response_products,
            )

        except Exception as e:
            print(f"❌ Error en [{self.supplier_name}] execute(): {e}")
            raise

    # ------------------------------------------------------------------ #
    #  Automatización                                                       #
    # ------------------------------------------------------------------ #

    def run_automation(
        self,
        email: str,
        password: str,
        csv_filename: str,
        po_data: Optional[PurchaseOrderDataModel] = None,
        po_items: Optional[List[Dict]] = None,
        on_item_ready=None,
        **kwargs,
    ) -> Optional[List[Dict]]:
        items = po_items or []
        if not items and po_data is not None:
            items = [
                {
                    "part_number": p.partNumber,
                    "qty": p.qty,
                    "mfrid": p.mfrid,
                    "mfrid_orig": p.mfrid_orig
                }
                for p in po_data.products
            ]
        return florida_outdoor_automation_playwright(
            username=email,
            password=password,
            po_items=items,
            on_item_ready=on_item_ready,
        )

    # ------------------------------------------------------------------ #
    #  Procesamiento de resultados                                          #
    # ------------------------------------------------------------------ #

    def process_results(
        self,
        scraped_data: List[Dict],
        po_data: PurchaseOrderDataModel,
    ) -> List[PurchaseOrderResponseProduct]:
        """
        Compara `your_price` (del carrito FOE) con `idealCost` de la PO.
        Tolerancia: 1 % del ideal_cost.

        Statuses de salida:
          CORRECT    → precio dentro de la tolerancia
          MISMATCH   → precio fuera de la tolerancia
          PART_ERROR → ítem inválido (order pad) o sin precio
        """
        ideal_costs: Dict[str, float] = {
            p.partNumber: p.idealCost for p in po_data.products
        }
        mfrid_map: Dict[str, str] = {
            p.partNumber: p.mfrid for p in po_data.products
        }
        mfrid_orig_map: Dict[str, str] = {
            p.partNumber: p.mfrid_orig for p in po_data.products
        }

        response_products: List[PurchaseOrderResponseProduct] = []

        for item in scraped_data:
            part_number: str = item.get("part_number", "")
            # Enriquecer mfrid: si viene del scraper usarlo, sino del mfrid_map
            scraped_mfrid = item.get("mfrid", "")
            item["mfrid"] = scraped_mfrid if scraped_mfrid else mfrid_map.get(part_number, "")
            # ✅ Asegurar que mfrid_orig se preserva desde PO
            if not item.get("mfrid_orig"):
                item["mfrid_orig"] = mfrid_orig_map.get(part_number, "")
            your_price: Optional[Decimal] = item.get("your_price")
            pre_status: str = item.get("status", "CORRECT")
            error_message: Optional[str] = item.get("error_message")
            pack_qty: Optional[int] = item.get("pack_qty")
            nla: Optional[str] = item.get("nla")
            ltl: Optional[str] = item.get("ltl")
            superseded_from: Optional[str] = item.get("superseded_from")
            cart_qty: int = item.get("qty", 0)

            ideal_cost = ideal_costs.get(part_number, 0.0)
            item["ideal_cost"] = ideal_cost
            price_float = float(your_price) if your_price is not None else 0.0

            # ── Statuses pre-asignados por el scraper ──────────────────────
            if pre_status == "PART_ERROR" and price_float == 0.0:
                # Ítem verdaderamente inválido: no encontrado o sin precio
                status = "PART_ERROR"
                print(f"  ❌ PART_ERROR [{part_number}]: {error_message}")

            elif pre_status == "SUPERSEDED":
                status = "SUPERSEDED"
                print(f"  🔄 SUPERSEDED [{part_number}] ← {superseded_from}")

            # ── Sin precio del carrito ─────────────────────────────────────
            elif price_float == 0.0:
                # Precio pendiente de scraping de carrito
                status = "CORRECT"
                print(f"  ⏳ SIN PRECIO (pendiente carrito): {part_number}")

            # ── Comparación de precios ─────────────────────────────────────
            elif price_float > 0 and ideal_cost > 0:
                difference = abs(ideal_cost - price_float)
                tolerance = ideal_cost * 0.01
                print(
                    f"  💵 [{part_number}] Ideal=${ideal_cost:.2f} | "
                    f"FOE=${price_float:.2f} | "
                    f"Diff=${difference:.2f} | Tol=${tolerance:.2f}"
                )

                if difference > tolerance:
                    status = "MISMATCH"
                    error_message = (
                        f"Price mismatch: Expected ${ideal_cost:.2f}, "
                        f"FOE ${price_float:.2f}"
                    )
                    if pack_qty:
                        error_message += f" (Pack item — min qty: {pack_qty})"
                    print(f"  ❌ MISMATCH: {part_number}")
                else:
                    status = "CORRECT"
                    notes = []
                    if pack_qty:
                        notes.append(f"Pack item — min qty: {pack_qty}")
                    if ltl:
                        notes.append("LTL shipment required")
                    error_message = " | ".join(notes) if notes else None
                    print(f"  ✅ CORRECT: {part_number}")

            else:
                status = "CORRECT"

            item["status"] = status

            # ✅ Asegurar que mfrid se propaga a la response
            final_mfrid = item.get("mfrid", "")
            if not final_mfrid:
                final_mfrid = mfrid_map.get(part_number, "")
            
            response_products.append(
                PurchaseOrderResponseProduct(
                    mfrid=final_mfrid,
                    partNumber=part_number,
                    qty=cart_qty,
                    idealCost=ideal_cost if ideal_cost > 0 else 0.0,
                    supplierPrice=price_float,
                    status=status,
                    nla=nla,
                    supersededFrom=superseded_from,
                    packQty=pack_qty,
                    ltl=ltl,
                )
            )

        # Nota: la inserción en BD se realiza en execute() ítem a ítem.
        # process_results() solo calcula statuses — no inserta directamente.
        return response_products
